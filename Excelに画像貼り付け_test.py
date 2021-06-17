#https://news.mynavi.jp/article/zeropython-42/
#Excelに画像貼り付け.py
#https://www.cresco.co.jp/blog/entry/11916/
import os
import glob
import imghdr
import openpyxl
import tkinter,tkinter.filedialog,pathlib,sys,io
from PIL import Image, ImageFilter
from datetime import datetime
import logging,traceback
import tkinter.messagebox

# 定数設定
#INPUT_IMG_DIR = '.\\test_images\\' # 貼り付ける画像を置いておくルートディレクトリ
SHEET_TITLE = '画像貼り付け' # シート名の設定
RESULT_FILE_NAME = '.\\result.xlsx' # 結果を保存するファイル名
L_LEN = 500 # 画像の最大幅 
# 変数
#max_height = [] # 各行の画像の高さの最大値を保持
 
 
def get_file_names(set_dir_name):
	"""
	ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
	"""
	file_names = os.listdir(set_dir_name)
	temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(os.path.join(set_dir_name, file_name))] # ファイルかどうかを判定
	return temp_full_file_names

def attach_img(dirs, set_column_idx):


	#"""
	#画像を呼び出して、Excelに貼り付け
	#"""
	set_row_idx = 1
	temp_count = 0
	ltemp_count = 1
	iDir=os.path.abspath(os.path.dirname(__file__))
	try:
		if os.path.exists(os.path.join(iDir,"temp")) == True:
			shutil.rmtree(os.path.join(iDir,"temp"))
		os.mkdir(os.path.join(iDir,"temp"),True)
		logging.basicConfig(filename='temp\logFile.txt',level=logging.INFO)
	except:
		tkinter.messagebox.showerror("初期化に失敗",traceback.format_exc())
	try:
		wb = openpyxl.Workbook()
		ws = wb.worksheets[0] # 1番目のシートを編集対象にする
		ws.title = SHEET_TITLE # 1番目のシートに名前を設定
		column_letter = ws.cell(row=set_row_idx, column=set_column_idx+1).column_letter # セルの行列番号から、そのセルの列番号の文字列を取得 colum=>colum_letter
		ws.cell(row=set_row_idx, column=set_column_idx).value = "No"
		ws.cell(row=set_row_idx, column=set_column_idx+1).value = "画像"
		ws.cell(row=set_row_idx, column=set_column_idx+2).value = "フルパス"
		ws.cell(row=set_row_idx, column=set_column_idx+3).value = "作成日時"
		ws.cell(row=set_row_idx, column=set_column_idx+4).value = "更新日時"
		ws.cell(row=set_row_idx, column=set_column_idx+5).value = "ファイルサイズ"
	# 各ディレクトリについて操作
		for set_dir_name in dirs:
			target_full_file_names = get_file_names(set_dir_name) # ファイル名取得
			#ws.cell(row=1, column=set_column_idx).value = set_dir_name # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
			max_width = 0 # 画像の幅の最大値を保持するための変数
			target_full_file_names.sort() # ファイル名でソート
			for target_file in target_full_file_names:
				if imghdr.what(target_file) != None: # 画像ファイルかどうかの判定
					img_temp = Image.open(target_file)
						# height と　width をくらべ、大きいほうをs_lenに代入
					height = img_temp.height
					width = img_temp.width
					if height > width:
						s_len = height
					else:
						s_len = width
					if L_LEN < s_len:
						height = int(height * L_LEN / s_len)
						width = int(width * L_LEN / s_len)
						img_temp=img_temp.resize((width,height))
						img_temp.save( os.path.join(os.path.join(iDir,"temp") , str(temp_count)+".png" ) )
						img = openpyxl.drawing.image.Image(os.path.join(os.path.join(iDir,"temp") , str(temp_count)+".png" ))
						temp_count +=1
					
					else:
						img = openpyxl.drawing.image.Image(target_file)
				# L_LENとs_lenをくらべ、L_LENが小さければL_LEN/s_lenで画像サイズを修正
				
					print('[' + str(ltemp_count) + ']' + target_file + 'を貼り付け')
	
					try:
			
					#height, width = size_img.shape[:2] 縦横はopenpyxl.drawing.image.Imageが持っている
						height = img.height
						width = img.width
						if max_width < width:
							max_width = width

						ws.row_dimensions[set_row_idx+1].height = height
						ws.column_dimensions[column_letter].width = max_width * 0.14
						cell_address = ws.cell(row=set_row_idx + 1, column=set_column_idx+1).coordinate # セルの行列番号から、そのセルの番地を取得
						img.anchor = cell_address
						ws.add_image(img) # シートに画像貼り付け
						ws.cell(row=set_row_idx+1, column=set_column_idx).value = ltemp_count
						ws.cell(row=set_row_idx+1, column=set_column_idx+2).value = target_file
						ws.cell(row=set_row_idx+1, column=set_column_idx+3).value = str(datetime.fromtimestamp(os.path.getctime(target_file)))
						ws.cell(row=set_row_idx+1, column=set_column_idx+4).value = str(datetime.fromtimestamp(os.path.getmtime(target_file)))
						ws.cell(row=set_row_idx+1, column=set_column_idx+5).value = str(os.path.getsize(target_file))
						set_row_idx += 1 #画像を張り付けた時だけ段を下げるようにする
						ltemp_count +=1
					except AttributeError as err:
						continue
					if set_row_idx > 500:
						set_row_idx = 1
						max_width = 0
						wb.save(RESULT_FILE_NAME[:-5]+str(ltemp_count)+'.xlsx')
						wb.close()
						wb = openpyxl.Workbook()
						ws = wb.worksheets[0] # 1番目のシートを編集対象にする
						ws.title = SHEET_TITLE # 1番目のシートに名前を設定
						ws.cell(row=set_row_idx, column=set_column_idx).value = "No"
						ws.cell(row=set_row_idx, column=set_column_idx+1).value = "画像"
						ws.cell(row=set_row_idx, column=set_column_idx+2).value = "フルパス"
						ws.cell(row=set_row_idx, column=set_column_idx+3).value = "作成日時"
						ws.cell(row=set_row_idx, column=set_column_idx+4).value = "アクセス日時"
						ws.cell(row=set_row_idx, column=set_column_idx+5).value = "ファイルサイズ"
						column_letter = ws.cell(row=set_row_idx, column=set_column_idx+1).column_letter # セルの行列番号から、そのセルの列番号の文字列を取得 colum=>colum_letter
					
		# ファイルへの書き込み
		wb.save(RESULT_FILE_NAME[:-5]+str(ltemp_count)+'.xlsx')
		tkinter.messagebox.showinfo("完了","作業完了しました。")
	except:
		tkinter.messagebox.showerror("エラー",traceback.format_exc())
		logging.error(traceback.format_exc())

root = tkinter.Tk()
root.withdraw()
fTyp =[("","*")]
iDir=os.path.abspath(os.path.dirname(__file__))

print("対象フォルダを選んでください")
try:
	dir=tkinter.filedialog.askdirectory(initialdir = iDir)

# 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
	dirs = glob.glob(os.path.join(dir, '**' + os.sep), recursive=True)
 
	set_column_idx = 1
except:
	tkinter.messagebox.showerror("エラー",traceback.format_exc())

attach_img(dirs, set_column_idx) # 画像貼り付け設定
#set_column_idx += 1 # 次の列へ・・・
